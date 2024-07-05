
from openpyxl import load_workbook
import pandas as pd
import os

def recalculate_excel(file_path):
    os.system(f"libreoffice --headless --convert-to xls {file_path}")
    os.system(f"rm {file_path}")
    name = file_path.split('.')
    print(f"{name[0]}.xls")
    os.system(f"libreoffice --headless --convert-to xlsx {name[0]}.xls")
    os.system(f"rm {name[0]}.xls")
    
def get_new_excel(input_file_name, output_file_name, *args, **kwargs):
    #load excel file
    workbook = load_workbook(filename=input_file_name)
    
    #open workbook
    sheet = workbook.worksheets[0]
    sheet["A1"] = 7
    sheet["A2"] = 7
    sheet["A3"] = 7
    
    #save the file
    workbook.save(filename=output_file_name)

    # Recalculer et sauvegarder le fichier
    recalculate_excel(output_file_name)

    # Le fichier recalculé sera sauvegardé dans le même répertoire avec le même nom
    print(f"Le fichier recalculé est sauvegardé")
    
    workbook2 = load_workbook(filename=output_file_name, data_only=True)
    sheets = workbook2.worksheets[1]
    print(sheets["A1"].value)
    
    os.system(f"rm {output_file_name}")
    
    
get_new_excel("test.xlsx", "output.xlsx")